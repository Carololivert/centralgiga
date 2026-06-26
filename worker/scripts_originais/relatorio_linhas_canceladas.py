#!/usr/bin/env python3
"""
relatorio_linhas_canceladas.py

Extrai os números de telefone (campo "Linhas") dos contratos de
"Telefone fixo ilimitado" CANCELADOS no SGP (GIGANET), via GET (sem navegador),
e gera um relatório Excel. Varre mês a mês.

Campos do relatório (confirmados pelo inspetor):
  tipo_pesquisa=0 (Situação Atual) | status=3 (Cancelado) | contrato=5 (Tel. fixo ilimitado)
  data_inicial / data_final (DD/MM/AAAA) | paginate_by=1000
  + dpb_token (escondido, puxado da própria página)
"""

import os
import re
import calendar
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

SGP_BASE   = "https://giganetwireless.sgp.net.br"
LOGIN_URL  = f"{SGP_BASE}/accounts/login"
REPORT_URL = f"{SGP_BASE}/admin/relatorios/contrato/status/"

# ---- Período a varrer (mês a mês) -----------------------------------------
ANO_INICIO, MES_INICIO = 2026, 1     # janeiro/2025
ANO_FIM,    MES_FIM    = 2026, 6    # dezembro/2025

# ---- Filtros (códigos confirmados) ----------------------------------------
TIPO_PESQUISA = "0"   # 0=Situação Atual | 2=Situação no Período (mude p/ "2" se vier vazio)
STATUS        = "3"   # Cancelado
CONTRATO      = "5"   # Telefone fixo ilimitado
PAGINATE_BY   = "1000"

SAIDA = "linhas_telefone_canceladas.xlsx"


def ssl_verify_enabled() -> bool:
    return os.getenv("SGP_VERIFY_SSL", "true").strip().lower() not in {"0", "false", "no", "off"}


if not ssl_verify_enabled():
    requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)


# =========================
# LOGIN (igual o seu)
# =========================
def login_sgp(session: requests.Session) -> None:
    session.headers.update({"User-Agent": "Mozilla/5.0", "Referer": LOGIN_URL})
    user, pwd = os.getenv("SGP_USER"), os.getenv("SGP_PASS")
    if not user or not pwd:
        raise Exception("SGP_USER / SGP_PASS não definidos no .env")
    v = ssl_verify_enabled()
    res = session.get(LOGIN_URL, verify=v)
    soup = BeautifulSoup(res.text, "html.parser")
    tok = soup.find("input", {"name": "csrfmiddlewaretoken"})
    payload = {"username": user, "password": pwd}
    if tok:
        payload["csrfmiddlewaretoken"] = tok["value"]
    session.post(LOGIN_URL, data=payload, allow_redirects=True, verify=v).raise_for_status()


# =========================
# PARÂMETROS
# =========================
def ultimo_dia(ano, mes):
    return calendar.monthrange(ano, mes)[1]


def meses_no_intervalo(ano_i, mes_i, ano_f, mes_f):
    y, m = ano_i, mes_i
    while (y, m) <= (ano_f, mes_f):
        yield y, m
        m += 1
        if m > 12:
            m, y = 1, y + 1


def obter_tokens(session: requests.Session) -> dict:
    """Pega os campos escondidos exigidos pelo formulário (dpb_token, csrf...)."""
    r = session.get(REPORT_URL, verify=ssl_verify_enabled())
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    tokens = {}
    for name in ("csrfmiddlewaretoken", "dpb_token"):
        el = soup.find("input", {"name": name})
        if el is not None:
            tokens[name] = el.get("value", "")
    return tokens


def build_params(ano: int, mes: int, tokens: dict) -> dict:
    p = {
        "tipo_pesquisa": TIPO_PESQUISA,
        "status": STATUS,
        "contrato": CONTRATO,
        "data_inicial": f"01/{mes:02d}/{ano}",
        "data_final": f"{ultimo_dia(ano, mes):02d}/{mes:02d}/{ano}",
        "paginate_by": PAGINATE_BY,
    }
    p.update(tokens)
    return p


# =========================
# PARSE DA TABELA
# =========================
def parse_contratos(html: str):
    soup = BeautifulSoup(html, "html.parser")
    table = soup.select_one("table#DataTables_Table_0")
    if table is None:
        for t in soup.find_all("table"):
            if "Observa" in t.get_text():
                table = t
                break
    if table is None:
        return []

    contratos, vistos = [], set()
    for tr in (table.select("tbody tr") or table.select("tr")):
        a = tr.find("a", href=True)
        if not a:
            continue
        label = a.get_text(strip=True)
        url = urljoin(SGP_BASE, a["href"])
        if label and url not in vistos:
            vistos.add(url)
            contratos.append((label, url))
    return contratos


# =========================
# EXTRAÇÃO DO "LINHAS"
# =========================
def extrair_linhas(session: requests.Session, url: str):
    r = session.get(url, verify=ssl_verify_enabled())
    r.raise_for_status()
    texto = BeautifulSoup(r.text, "html.parser").get_text("\n")

    numeros = []
    m = re.search(r"Linhas?\b[\s:]*([\s\S]{0,80})", texto, re.I)
    if m:
        trecho = re.split(r"(Situa\w*|Plano|Contrato\s*Tipo|Tipo\s*Pessoa)", m.group(1))[0]
        for chunk in re.split(r"[,;/\n]", trecho):
            d = re.sub(r"\D", "", chunk)
            if 8 <= len(d) <= 13:
                numeros.append(d)
    return numeros


# =========================
# EXCEL
# =========================
def salvar_excel(rows, caminho):
    wb = Workbook()
    ws = wb.active
    ws.title = "Linhas Canceladas"
    ws.append(["Mês", "Contrato", "Cliente", "Linha(s)"])
    azul = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.fill = azul
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")
    cinza = PatternFill("solid", fgColor="F2F2F2")
    for idx, r in enumerate(rows, start=2):
        ws.append(r)
        if idx % 2 == 0:
            for c in ws[idx]:
                c.fill = cinza

    # ---- destaque: NÚMEROS LIVRES (todas as linhas dos cancelados, sem repetir) ----
    livres, vistos = [], set()
    for r in rows:
        for n in str(r[3]).split(";"):
            n = n.strip()
            if n and n not in vistos:
                vistos.add(n)
                livres.append(n)

    verde       = PatternFill("solid", fgColor="00B050")
    verde_claro = PatternFill("solid", fgColor="C6EFCE")
    ini = ws.max_row + 2  # deixa uma linha em branco antes
    cab = ws.cell(row=ini, column=1, value=f"NÚMEROS LIVRES ({len(livres)})")
    cab.fill = verde
    cab.font = Font(color="FFFFFF", bold=True, size=12)
    ws.merge_cells(start_row=ini, start_column=1, end_row=ini, end_column=4)
    for i, num in enumerate(livres, start=ini + 1):
        cel = ws.cell(row=i, column=1, value=num)
        cel.fill = verde_claro
        cel.font = Font(bold=True)

    for col, w in zip("ABCD", [16, 14, 34, 26]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(caminho)


# =========================
# MAIN
# =========================
def main():
    session = requests.Session()
    print("Logando no SGP...")
    login_sgp(session)

    tokens = obter_tokens(session)
    rows = []
    for ano, mes in meses_no_intervalo(ANO_INICIO, MES_INICIO, ANO_FIM, MES_FIM):
        print(f"\n=== {mes:02d}/{ano} ===")
        r = session.get(REPORT_URL, params=build_params(ano, mes, tokens), verify=ssl_verify_enabled())
        r.raise_for_status()

        contratos = parse_contratos(r.text)
        print(f"  {len(contratos)} contrato(s) cancelado(s)")

        for label, url in contratos:
            numeros = extrair_linhas(session, url)
            if numeros:
                num, _, nome = label.partition(" - ")
                rows.append([f"{mes:02d}/{ano}", num.strip(), nome.strip(), "; ".join(numeros)])
                print(f"    OK  {label}  ->  {', '.join(numeros)}")
            else:
                print(f"    --  {label}  (sem linha)")

    salvar_excel(rows, SAIDA)
    print(f"\nPronto! {len(rows)} linha(s) -> {SAIDA}")


if __name__ == "__main__":
    main()